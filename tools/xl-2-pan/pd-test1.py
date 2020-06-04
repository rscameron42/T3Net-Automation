from openpyxl import load_workbook
import pandas as pd

def get_all_tables(filename):
    """ Get all tables from a given workbook. Returns a dictionary of tables. 
        Requires a filename, which includes the file path and filename. """

    # Load the workbook, from the filename
    wb = load_workbook(filename=file, read_only=False, keep_vba=False, data_only=True, keep_links=False)

    # Initialize the dictionary of tables
    tables_dict = {}

    # Go through each worksheet in the workbook
    for ws_name in wb.sheetnames:
        print("")
        print(f"worksheet name: {ws_name}")
        ws = wb[ws_name]
        print(f"tables in worksheet: {len(ws._tables)}")

        # Get each table in the worksheet
        for tbl in ws._tables:
            print(f"\ttable name: {tbl.name}")
            # First, add some info about the table to the dictionary
            tables_dict[tbl.name] = {
                'table_name': tbl.name,
                'worksheet': ws_name,
                'num_cols': len(tbl.tableColumns),
                'table_range': tbl.ref}

            # Grab the 'data' from the table
            data = ws[tbl.ref]
            
            # Now convert the table 'data' to a Pandas DataFrame
            # First get a list of all rows, including the first header row
            rows_list = []
            for row in data:
                # Get a list of all columns in each row
                cols = []
                for col in row:
                    cols.append(col.value)
                rows_list.append(cols)

            # Create a pandas dataframe from the rows_list. 
            # The first row is the column names
            df = pd.DataFrame(data=rows_list[1:], index=None, columns=rows_list[0])
            
            # Add the dataframe to the dictionary of tables
            tables_dict[tbl.name]['dataframe'] = df

    return tables_dict

# Define levels of headers for readability and CLI code seperation
def header1(msg):
    print(f"\n{'#' * 90}\n####\n#### {msg}\n####\n{'#' * 90}")
def header2(msg):
    print(f"\n{'-' * 90}\n---- {msg}\n{'-' * 90}")

def addresses_cli():
    # Process Addresses Table
    header1("Print the Addresses dataframe, then iterate through the data to generate CLI commands")
    print(addr_df)
    for _, row in addr_df.iterrows():
        header2(f"Add / Modify {row['name']}")
        corecmd = " address " + row["name"] + " "
        if row["device-group"]:
            corecmd = " device-group " + row["device-group"] + corecmd
        print(f"set{corecmd}{row['type']} {row['address']}")
        if row["tag_mod"] == "replace": 
            print(f"delete{corecmd}tag" ) 
        if row["tag"]:
            print(f"set{corecmd}tag [ {row['tag']} ]")
        if row["description"]:
            print(f"set{corecmd}description \"{row['description']}\"")

def addressgroups_cli():
    # Process Address Groups Table
    header1("Print the Address Groups dataframe, then iterate through the data to generate CLI commands")
    print(addrgrp_df)
    for _, row in addrgrp_df.iterrows():
        header2(f"Add / Modify {row['name']}")
        corecmd = " address-group " + row["name"] + " "
        if row["device-group"]:
            corecmd = " device-group " + row["device-group"] + corecmd
        if row["type"] == "dynamic":
            print(f"set{corecmd}{row['type']} filter \"{row['filter']}\"")
        else:
            if row["members_mod"] == "replace":
                 print(f"delete{corecmd}members" )
            print(f"set{corecmd}members [ {row['members']} ]")
        if row["tag_mod"] == "replace": 
            print(f"delete{corecmd}tag" ) 
        if row["tag"]:
            print(f"set{corecmd}tag [ {row['tag']} ]")
        if row["description"]:
            print(f"set{corecmd}description \"{row['description']}\"")

def tags_cli():
    # Process Tags Table
    header1("Print the Tags dataframe, then iterate through the data to generate CLI commands")
    print(tag_df)
    for _, row in tag_df.iterrows():
        header2(f"Add / Modify {row['name']}")
        corecmd = " tag " + row["name"] + " "
        if row["device-group"]:
            corecmd = " device-group " + row["device-group"] + corecmd
        if row["color"]:
            corecmd = corecmd + "color " + row["color"] + " "
        if row["comments"]:
            corecmd = corecmd + "comments \"" + row["comments"] + "\""
        print(f"set{corecmd}")

# File location:
file = r"wb23.xlsx"

# Run the function to return a dictionary of all tables in the Excel workbook
tables_dict = get_all_tables(filename=file)

# Create dataframes for each named table in the Excel workbook
security_df =   pd.DataFrame(data=tables_dict['SecurityPolicy']['dataframe'])
addr_df     =   pd.DataFrame(data=tables_dict['Addresses']['dataframe'])
addrgrp_df  =   pd.DataFrame(data=tables_dict['AddressGroups']['dataframe'])
tag_df      =   pd.DataFrame(data=tables_dict['Tags']['dataframe'])
service_df  =   pd.DataFrame(data=tables_dict['Services']['dataframe'])
svcgrp_df   =   pd.DataFrame(data=tables_dict['ServiceGroups']['dataframe'])

tags_cli()
addresses_cli()
addressgroups_cli()
