from openpyxl import load_workbook
import pandas as pd

def get_all_tables(filename):
    """ Get all tables from a given workbook. Returns a dictionary of tables. 
        Requires a filename, which includes the file path and filename. """

    # Load the workbook, from the filename
    wb = load_workbook(filename=file, read_only=False, keep_vba=False, data_only=True, keep_links=False)

    # Initialize the dictionary of tables
    tables_dict = {}

    # Go through each worksheet in the workbook
    for ws_name in wb.sheetnames:
        print("")
        print(f"worksheet name: {ws_name}")
        ws = wb[ws_name]
        print(f"tables in worksheet: {len(ws._tables)}")

        # Get each table in the worksheet
        for tbl in ws._tables:
            print(f"table name: {tbl.name}")
            # First, add some info about the table to the dictionary
            tables_dict[tbl.name] = {
                    'table_name': tbl.name,
                    'worksheet': ws_name,
                    'num_cols': len(tbl.tableColumns),
                    'table_range': tbl.ref}

            # Grab the 'data' from the table
            data = ws[tbl.ref]

            # Now convert the table 'data' to a Pandas DataFrame
            # First get a list of all rows, including the first header row
            rows_list = []
            for row in data:
                # Get a list of all columns in each row
                cols = []
                for col in row:
                    cols.append(col.value)
                rows_list.append(cols)

            # Create a pandas dataframe from the rows_list. 
            # The first row is the column names
            df = pd.DataFrame(data=rows_list[1:], index=None, columns=rows_list[0])

            # Add the dataframe to the dictionary of tables
            tables_dict[tbl.name]['dataframe'] = df

    return tables_dict

# File location:
file = r"/Users/rcameron/Documents/pan-ss2fw-tools/wb21.xlsx"

# Run the function to return a dictionary of all tables in the Excel workbook
tables_dict = get_all_tables(filename=file)

print(tables_dict)