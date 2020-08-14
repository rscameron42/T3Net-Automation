#!/usr/bin/env python3

# Convert Excel data to Palo Alto CLI configuration commands

""" 
Pull data from multiple tables in an Excel file and use it to add, modify 
or delete configuration data from a Palo Alto Network's firewall or  
Panorama managment server.
Current version (1.0.x):
Output is only CLI commands used to past into a config-level CLI session.
Source data can be:
Policies [ Security ]
Objects [ Addresses Address-Groups Services Service-Groups Tags ]
"""

__author__ = "Ron Cameron"
__copyright__ = "Copyright 2020, Remove kp-all Policy Cleanup"
__credits__ = ["Ron Cameron"]

__license__ = '{license}'
__version__ = "1.0.4"
__maintainer__ = "Ron Cameron"
__email__ = "rscameron@gmail.com"
__status__ = '{dev_status}'

# Futures
#from __future__ import print_function
# […]

# Built-in/Generic Imports
import sys
import os
import pandas as pd

# Libs
from openpyxl import load_workbook
# […]
# […]

# Own modules
#from {path} import {class}
# […]

def get_all_tables(filename): # REVIEW - using just the dictionary and not Pandas dataframe
    """ Get all tables from a given workbook. Returns a dictionary of tables. 
        Requires a filename, which includes the file path and filename. """

    # Load the workbook, from the filename
    wb = load_workbook(filename=file, read_only=False, keep_vba=False, data_only=True, keep_links=False)

    # Initialize the dictionary of tables
    tables_dict = {}

    # Go through each worksheet in the workbook
    for ws_name in wb.sheetnames:
        print("")
        print(f"worksheet name: {ws_name}")
        ws = wb[ws_name]
        print(f"tables in worksheet: {len(ws._tables)}")

        # Get each table in the worksheet
        for tbl in ws._tables:
            print(f"\ttable name: {tbl.name}")
            # First, add some info about the table to the dictionary
            tables_dict[tbl.name] = {
                'table_name': tbl.name,
                'worksheet': ws_name,
                'num_cols': len(tbl.tableColumns),
                'table_range': tbl.ref}

            # Grab the 'data' from the table
            data = ws[tbl.ref]
            
            # Now convert the table 'data' to a Pandas DataFrame
            # First get a list of all rows, including the first header row
            rows_list = []
            for row in data:
                # Get a list of all columns in each row
                cols = []
                for col in row:
                    cols.append(col.value)
                rows_list.append(cols)

            # Create a pandas dataframe from the rows_list. 
            # The first row is the column names
            df = pd.DataFrame(data=rows_list[1:], index=None, columns=rows_list[0])
            
            # Add the dataframe to the dictionary of tables
            tables_dict[tbl.name]['dataframe'] = df
    return tables_dict

# Define levels of headers for readability and CLI code seperation
def header1(msg):
    print(f"\n{'#' * 90}\n####\n#### {msg}\n####\n{'#' * 90}")
def header2(msg):
    print(f"\n{'-' * 90}\n---- {msg}\n{'-' * 90}")

def security_cli():  # TODO - Finish Security Policy CLI Function  
    # Process Addresses Table
    header1("Print the SecurityPolicy dataframe")
    print(security_df)
    for _, row in security_df.iterrows():
        header2(f"{row['operation']} {row['device-group']} {row['dg_rulebase']} Security Policy: {row['name']}")
        corecmd = "rulebase security rules \"" + row["name"] + "\""
        clonecmd = "rulebase security rules \"" + str(row["src_rule"]) + "\" to \"" + row["name"] + "\""
        if row["device-group"]:
            corecmd = "device-group " + row["device-group"] + " " + row["dg_rulebase"] + "-" + corecmd
            clonecmd = "device-group " + row["device-group"] + " " + row["dg_rulebase"] + "-" + clonecmd
        if row["operation"] == "delete":
            print(f"delete {corecmd}")
            continue
        elif (row["operation"] == "copy") and (row["src_rule"]):
            print(f"copy {clonecmd}")
            if (row["relation_to"]):
                print(f"move {corecmd} {row['position']} \"{row['relation_to']}\"")
        elif (row["operation"] == "create"):
            print(f"set {corecmd} disabled no")
            if (row["relation_to"]):
                print(f"move {corecmd} {row['position']} \"{row['relation_to']}\"")
        if row["disabled"]:
            print(f"set {corecmd} disabled {row['disabled']}")
        if row["tag_mod"] == "replace": 
            print(f"delete {corecmd} tag" ) 
        if row["tag"]:
            print(f"set {corecmd} tag [ {row['tag']} ]")
        if row["from_mod"] == "replace": 
            print(f"delete {corecmd} from" ) 
        if row["from"]:
            print(f"set {corecmd} from [ {row['from']} ]")
        if row["source_mod"] == "replace": 
            print(f"delete {corecmd} source" ) 
        if row["source"]:
            print(f"set {corecmd} source [ {row['source']} ]")
        if row["to_mod"] == "replace": 
            print(f"delete {corecmd} to" ) 
        if row["to"]:
            print(f"set {corecmd} to [ {row['to']} ]")
        if row["destination_mod"] == "replace": 
            print(f"delete {corecmd} destination" ) 
        if row["destination"]:
            print(f"set {corecmd} destination [ {row['destination']} ]")
        if row["app_mod"] == "replace": 
            print(f"delete {corecmd} application" ) 
        if row["application"]:
            print(f"set {corecmd} application [ {row['application']} ]")
        if row["service_mod"] == "replace": 
            print(f"delete {corecmd} service" ) 
        if row["service"]:
            print(f"set {corecmd} service [ {row['service']} ]")
        if row["action"]:
            print(f"set {corecmd} action {row['action']}")
        if row["description"]:
            print(f"set {corecmd} description \"{row['description']}\"")

def addresses_cli():
    # Process Addresses Table
    #header1("Print the Addresses dataframe")
    #print(addr_df)
    for _, row in addr_df.iterrows():
        header2(f"Add / Modify Address: {row['name']}")
        corecmd = "address " + row["name"]
        if row["device-group"]:
            corecmd = "device-group " + row["device-group"] + " " + corecmd
        print(f"set {corecmd} {row['type']} {row['address']}")
        if row["tag_mod"] == "replace": 
            print(f"delete {corecmd} tag" ) 
        if row["tag"]:
            print(f"set {corecmd} tag [ {row['tag']} ]")
        if row["description"]:
            print(f"set {corecmd} description \"{row['description']}\"")

def addressgroups_cli():
    # Process Address Groups Table
    #header1("Print the Address Groups dataframe")
    print(addrgrp_df)
    for _, row in addrgrp_df.iterrows():
        header2(f"Add / Modify Address Group: {row['name']}")
        corecmd = "address-group " + row["name"]
        if row["device-group"]:
            corecmd = "device-group " + row["device-group"] + " " + corecmd
        if (row["type"] == "dynamic") and (row["filter"]):
            print(f"set {corecmd} {row['type']} filter \"\'{row['filter']}\"")
        elif row["type"] == "static":
            if row["members_mod"] == "replace":
                 print(f"delete {corecmd} members" )
            print(f"set {corecmd} members [ {row['members']} ]")
        if row["tag_mod"] == "replace": 
            print(f"delete {corecmd} tag" ) 
        if row["tag"]:
            print(f"set {corecmd} tag [ {row['tag']} ]")
        if row["description"]:
            print(f"set {corecmd} description \"{row['description']}\"")

def tags_cli():
    # Process Tags Table
    #header1("Print the Tags dataframe")
    #print(tag_df)
    for _, row in tag_df.iterrows():
        header2(f"Add / Modify Tag: {row['name']}")
        corecmd = "tag " + row["name"]
        if row["device-group"]:
            corecmd = "device-group " + row["device-group"] + " " + corecmd
        if row["color"]:
            corecmd = corecmd + "color " + row["color"]
        if row["comments"]:
            corecmd = corecmd + " comments \"" + row["comments"] + "\""
        print(f"set {corecmd}")

def services_cli():
    # Process Services Table
    #header1("Print the Services dataframe")
    #print(service_df)
    for _, row in service_df.iterrows():
        header2(f"Add / Modify Service: {row['name']}")
        corecmd = "service " + row["name"]
        if row["operation"] == "move":
            print(f"rename device-group {row['device-group']} service {row['source']} to {row['name']}")
            continue
        if row["device-group"]:
            corecmd = "device-group " + row["device-group"] + " " + corecmd
        if row["port"]:
            svcproto = "protocol " + row["protocol"] + " port " + str(int(row["port"]))
        if row["source-port"]:
            svcproto = svcproto + " source-port " + row["source-port"]
        print(f"set {corecmd} {svcproto}")
        if row["tag_mod"] == "replace": 
            print(f"delete {corecmd} tag" ) 
        if row["tag"]:
            print(f"set {corecmd} tag [ {row['tag']} ]")
        if row["description"]:
            print(f"set {corecmd} description \"{row['description']}\"")

def servicegroups_cli():
    # Process Service Groups Table
    #header1("Print the ServiceGroups dataframe")
    #print(svcgrp_df)
    for _, row in svcgrp_df.iterrows():
        header2(f"Add / Modify Service Group: {row['name']}")
        corecmd = "service-group " + row["name"]
        if row["device-group"]:
            corecmd = "device-group " + row["device-group"] + " " + corecmd
        if row["tag_mod"] == "replace": 
            print(f"delete {corecmd} tag" )
        if row["tag"]:
            print(f"set {corecmd} tag [ {row['tag']} ]")
        if row["members_mod"] == "replace": 
            print(f"delete {corecmd} members")
        if row["members"]:
            print(f"set {corecmd} members [ {row['members']} ]")

# File location:
if sys.argv[1]:
    file = (sys.argv[1])
    #print(f"file is: {file}")
else:
#file = r"xl2pan - test-data.xlsx"
    file = r"/Users/rcameron/Desktop/Week25/xl2pan-week25.xlsx"

# Run the function to return a dictionary of all tables in the Excel workbook
tables_dict = get_all_tables(filename=file)

# Create dataframes for each named table in the Excel workbook
security_df =   pd.DataFrame(data=tables_dict['SecurityPolicy']['dataframe'])
addr_df     =   pd.DataFrame(data=tables_dict['Addresses']['dataframe'])
addrgrp_df  =   pd.DataFrame(data=tables_dict['AddressGroups']['dataframe'])
tag_df      =   pd.DataFrame(data=tables_dict['Tags']['dataframe'])
service_df  =   pd.DataFrame(data=tables_dict['Services']['dataframe'])
svcgrp_df   =   pd.DataFrame(data=tables_dict['ServiceGroups']['dataframe'])

# Debug argument list
print ('Number of arguments:', len(sys.argv), 'arguments.')
print ('Argument List:', str(sys.argv))

tags_cli()
addresses_cli()
addressgroups_cli()
#services_cli()
#servicegroups_cli()
security_cli()
